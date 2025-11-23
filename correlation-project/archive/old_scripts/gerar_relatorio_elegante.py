"""
Gera relatório elegante para análise dos casos validados
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def criar_relatorio_analistas():
    print("=" * 80)
    print("GERANDO RELATÓRIO PARA ANALISTAS")
    print("=" * 80)
    
    # Carrega dados validados
    arquivo_validacao = Path('output/validacao_qwen3_otimizada.xlsx')
    
    if not arquivo_validacao.exists():
        print("❌ Execute primeiro: python validar_qwen3_otimizado.py")
        return
    
    print(f"\n📂 Carregando validações de: {arquivo_validacao}")
    df_todas = pd.read_excel(arquivo_validacao, sheet_name='Todas Validações')
    df_confirmadas = pd.read_excel(arquivo_validacao, sheet_name='CONFIRMADAS')
    
    # Prepara relatório consolidado
    relatorio = []
    
    for idx, row in df_confirmadas.iterrows():
        caso = {
            '🔍 Status': '✅ CONFIRMADO',
            'Nº': idx + 1,
            'Nome Completo': row['nome'],
            'Data Nascimento': row['data_nascimento'],
            'Nome da Mãe': row['nome_mae'],
            'Nome do Pai': row['nome_pai'],
            'RG': row['rg'],
            '🧠 Transtorno Psiquiátrico': row.get('tem_transtorno_psiquiatrico', 'Não informado'),
            '📋 Tipo Transtorno': row.get('tipo_transtorno', 'N/A'),
            
            '📋 BO Desaparecimento': row['bo_desaparecimento'],
            '⚰️ BO Morte/Cadáver': row['bo_morte'],
            'Tipo Morte': row['tipo_morte'],
            '⏱️ Dias Entre Eventos': row['dias_entre'],
            'Força Correlação': row['forca_correlacao'],
            
            '🤖 Confiança IA (%)': row['confianca'],
            '✅ Data Nasc OK': 'SIM' if row['data_nasc_ok'] else 'NÃO',
            '✅ Nome Mãe OK': 'SIM' if row['mae_ok'] else 'NÃO',
            '✅ Nome Pai OK': 'SIM' if row['pai_ok'] else 'NÃO',
            '✅ RG OK': 'SIM' if row['rg_ok'] else 'NÃO',
            
            '📝 Análise da IA': row['justificativa'],
            
            '👮 Análise do Perito': '',
            '📅 Data Análise': '',
            '✍️ Nome do Analista': '',
            '💭 Observações': ''
        }
        
        relatorio.append(caso)
    
    df_relatorio = pd.DataFrame(relatorio)
    
    # Salva Excel base
    arquivo_saida = Path('output/RELATORIO_ANALISE_CORRELACOES.xlsx')
    print(f"\n💾 Gerando relatório: {arquivo_saida}")
    
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        # Aba principal - PARA ANÁLISE
        df_relatorio.to_excel(writer, sheet_name='CASOS PARA ANÁLISE', index=False)
        
        # Aba resumo executivo
        resumo = {
            'Métrica': [
                'Total de Casos Validados pela IA',
                'Casos CONFIRMADOS (alta confiança)',
                'Confiança Média da IA',
                'Casos Críticos (0-7 dias)',
                'Casos Fortes (0-30 dias)',
                '',
                'Status Análise Humana',
                'Casos Analisados',
                'Casos Pendentes'
            ],
            'Valor': [
                len(df_confirmadas),
                len(df_confirmadas),
                f"{df_confirmadas['confianca'].mean():.1f}%",
                len(df_confirmadas[df_confirmadas['dias_entre'] <= 7]),
                len(df_confirmadas[df_confirmadas['forca_correlacao'] == 'FORTE']),
                '',
                '',
                '0',
                str(len(df_confirmadas))
            ]
        }
        
        pd.DataFrame(resumo).to_excel(writer, sheet_name='RESUMO EXECUTIVO', index=False)
        
        # Aba com dados completos de identificação
        df_dados_id = df_confirmadas[[
            'nome', 'data_nascimento', 'nome_mae', 'nome_pai', 'rg',
            'bo_desaparecimento', 'bo_morte', 'dias_entre', 'tipo_morte',
            'confianca', 'justificativa'
        ]].copy()
        
        df_dados_id.columns = [
            'Nome', 'Data Nascimento', 'Nome Mãe', 'Nome Pai', 'RG',
            'BO Desaparecimento', 'BO Morte', 'Dias', 'Tipo',
            'Confiança IA', 'Justificativa IA'
        ]
        
        df_dados_id.to_excel(writer, sheet_name='DADOS IDENTIFICAÇÃO', index=False)
    
    # Aplica formatação elegante
    print("   ✨ Aplicando formatação profissional...")
    formatar_excel(arquivo_saida)
    
    print("\n" + "=" * 80)
    print("✅ RELATÓRIO GERADO COM SUCESSO!")
    print("=" * 80)
    print(f"\n📊 Arquivo: {arquivo_saida}")
    print(f"\n📋 Conteúdo:")
    print(f"   • Aba 'CASOS PARA ANÁLISE': {len(df_relatorio)} casos com campos para análise humana")
    print(f"   • Aba 'RESUMO EXECUTIVO': Estatísticas gerais")
    print(f"   • Aba 'DADOS IDENTIFICAÇÃO': Dados completos para conferência")
    print(f"\n✅ Campos para preenchimento pelo analista:")
    print(f"   • 👮 Análise do Perito")
    print(f"   • 📅 Data Análise")
    print(f"   • ✍️ Nome do Analista")
    print(f"   • 💭 Observações")


def formatar_excel(arquivo):
    """Aplica formatação profissional ao Excel"""
    
    wb = load_workbook(arquivo)
    
    # Formata aba principal
    ws = wb['CASOS PARA ANÁLISE']
    
    # Cores
    cor_cabecalho = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cor_confirmado = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    cor_forte = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    cor_critico = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Fonte cabeçalho
    fonte_cabecalho = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    fonte_normal = Font(name='Arial', size=10)
    
    # Alinhamento
    alinhamento_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Bordas
    borda_fina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formata cabeçalho
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.fill = cor_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = alinhamento_centro
        cell.border = borda_fina
    
    # Formata dados
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = fonte_normal
            cell.border = borda_fina
            
            # Alinhamento
            if col in [1, 2, 11, 12, 13, 14, 15, 16, 17]:  # Colunas de texto curto
                cell.alignment = alinhamento_centro
            else:
                cell.alignment = alinhamento_esquerda
            
            # Destaque status
            if col == 1:  # Status
                cell.fill = cor_confirmado
            
            # Destaque dias críticos
            if col == 11:  # Dias entre eventos
                try:
                    dias = int(cell.value)
                    if dias <= 7:
                        cell.fill = cor_critico
                    elif dias <= 30:
                        cell.fill = cor_forte
                except:
                    pass
    
    # Ajusta largura das colunas
    larguras = {
        'A': 15,  # Status
        'B': 5,   # Nº
        'C': 35,  # Nome
        'D': 15,  # Data Nasc
        'E': 35,  # Nome Mãe
        'F': 35,  # Nome Pai
        'G': 15,  # RG
        'H': 20,  # BO Desap
        'I': 20,  # BO Morte
        'J': 12,  # Tipo
        'K': 8,   # Dias
        'L': 10,  # Força
        'M': 10,  # Confiança
        'N': 10,  # Data OK
        'O': 10,  # Mãe OK
        'P': 10,  # Pai OK
        'Q': 10,  # RG OK
        'R': 60,  # Análise IA
        'S': 60,  # Análise Perito
        'T': 15,  # Data
        'U': 25,  # Analista
        'V': 60   # Observações
    }
    
    for col_letter, width in larguras.items():
        ws.column_dimensions[col_letter].width = width
    
    # Congela primeira linha
    ws.freeze_panes = 'A2'
    
    # Formata aba RESUMO EXECUTIVO
    if 'RESUMO EXECUTIVO' in wb.sheetnames:
        ws_resumo = wb['RESUMO EXECUTIVO']
        
        # Cabeçalho
        for col in range(1, 3):
            cell = ws_resumo.cell(1, col)
            cell.fill = cor_cabecalho
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_centro
        
        # Dados
        for row in range(2, ws_resumo.max_row + 1):
            for col in range(1, 3):
                cell = ws_resumo.cell(row, col)
                cell.font = Font(name='Arial', size=11)
                cell.alignment = alinhamento_esquerda
                
                # Destaque métricas importantes
                if row <= 4:
                    cell.font = Font(name='Arial', size=11, bold=True)
        
        ws_resumo.column_dimensions['A'].width = 40
        ws_resumo.column_dimensions['B'].width = 20
    
    # Formata aba DADOS IDENTIFICAÇÃO
    if 'DADOS IDENTIFICAÇÃO' in wb.sheetnames:
        ws_dados = wb['DADOS IDENTIFICAÇÃO']
        
        # Cabeçalho
        for col in range(1, ws_dados.max_column + 1):
            cell = ws_dados.cell(1, col)
            cell.fill = cor_cabecalho
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_centro
        
        # Auto-ajusta colunas
        for col in range(1, ws_dados.max_column + 1):
            ws_dados.column_dimensions[get_column_letter(col)].width = 25
        
        # Justificativa mais larga
        ws_dados.column_dimensions['K'].width = 80
    
    wb.save(arquivo)


if __name__ == "__main__":
    criar_relatorio_analistas()
