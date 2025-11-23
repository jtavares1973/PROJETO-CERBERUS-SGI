"""Módulo para exportação elegante de dados em XLSX"""
import pandas as pd
from pathlib import Path
from typing import Optional
import re


def sanitizar_para_excel(valor):
    """Remove caracteres inválidos e lixo para Excel."""
    if pd.isna(valor):
        return ""
    
    if not isinstance(valor, str):
        return valor
    
    # 1. Tentar corrigir encoding UTF-8 mal-interpretado como Latin-1
    try:
        # Se foi lido como Latin-1 mas é UTF-8, reconverter
        if any(char in valor for char in ['Ã', 'Â', 'Ê', 'Ç', 'É']):
            try:
                valor = valor.encode('latin-1').decode('utf-8', errors='ignore')
            except (UnicodeDecodeError, UnicodeEncodeError):
                pass
    except:
        pass
    
    # 2. Mapeamento completo de caracteres UTF-8 mal-interpretados
    substituicoes_utf8 = {
        # Vogais com til
        'Ã£': 'ã', 'Ã¡': 'á', 'Ã¢': 'â', 'Ã ': 'à', 'Ãµ': 'õ', 'Ã³': 'ó', 'Ã´': 'ô',
        'Ã©': 'é', 'Ãª': 'ê', 'Ã­': 'í', 'Ãº': 'ú', 'Ã§': 'ç',
        # Maiúsculas
        'Ã': 'Ã', 'Ã‰': 'É', 'ÃŠ': 'Ê', 'Ã"': 'Ó', 'Ã‡': 'Ç',
        # Casos especiais
        'Â': '', 'Ã\x83': 'Ã', 'Ã\x82': 'Â',
    }
    
    for errado, correto in substituicoes_utf8.items():
        valor = valor.replace(errado, correto)
    
    # 3. Remove caracteres de controle (0x00-0x1F exceto tab, newline, CR)
    valor = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', valor)
    
    # 4. Remove BOM e zero-width characters
    valor = re.sub(r'[\uFEFF\u200B-\u200D\uFFFE\uFFFF]', '', valor)
    
    # 5. Substitui caracteres Latin-1 problemáticos remanescentes
    substituicoes_latin1 = {
        '\x82': 'é', '\x83': 'á', '\x84': 'â', '\x85': 'à', '\x86': 'ã',
        '\x87': 'ç', '\x88': 'ê', '\x89': 'é', '\x8a': 'ê', '\x8b': 'í',
        '\x8c': 'ó', '\x8d': 'ô', '\x8e': 'õ', '\x8f': 'ú', '\x90': 'É',
        '\x91': 'Á', '\x92': 'Â', '\x93': 'À', '\x94': 'Ã', '\x95': 'Ê',
        '\x96': 'Í', '\x97': 'Ó', '\x98': 'Ô', '\x99': 'Õ', '\x9a': 'Ú',
        '\x9b': 'ü', '\x9c': 'Ü', '\x9d': 'ñ', '\x9e': 'Ñ',
    }
    
    for char_antigo, char_novo in substituicoes_latin1.items():
        valor = valor.replace(char_antigo, char_novo)
    
    # 6. Remove APENAS caracteres de controle problemáticos, preserva acentos
    # Não usar isprintable() pois remove acentos válidos
    caracteres_proibidos = set(range(0x00, 0x09)) | set(range(0x0E, 0x20)) | {0x7F}
    valor = ''.join(c for c in valor if ord(c) not in caracteres_proibidos)
    
    # 7. Limpa múltiplos espaços e espaços estranhos
    valor = re.sub(r'\s+', ' ', valor).strip()
    
    # 8. Limita tamanho (Excel tem limite de 32767 caracteres por célula)
    if len(valor) > 32000:
        valor = valor[:32000] + "..."
    
    return valor


def exportar_excel_formatado(
    df: pd.DataFrame,
    caminho_saida: str,
    nome_planilha: str = "Dados",
    incluir_filtros: bool = True,
    congelar_paineis: bool = True,
    largura_auto: bool = True
) -> None:
    """
    Exporta DataFrame para Excel com formatação profissional.
    
    Args:
        df: DataFrame a exportar
        caminho_saida: Caminho do arquivo .xlsx
        nome_planilha: Nome da planilha
        incluir_filtros: Se True, adiciona filtros automáticos
        congelar_paineis: Se True, congela primeira linha
        largura_auto: Se True, ajusta largura das colunas
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Criar arquivo Excel
        caminho = Path(caminho_saida)
        caminho.parent.mkdir(parents=True, exist_ok=True)
        
        # Sanitizar DataFrame antes de salvar
        df_limpo = df.copy()
        for col in df_limpo.columns:
            if df_limpo[col].dtype == 'object':
                df_limpo[col] = df_limpo[col].apply(sanitizar_para_excel)
        
        # Salvar DataFrame
        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            df_limpo.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        # Carregar para formatação
        wb = load_workbook(caminho)
        ws = wb[nome_planilha]
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Borda
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Formatação do cabeçalho
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Formatação das células de dados
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Zebrado (linhas alternadas)
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajustar largura das colunas
        if largura_auto:
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                
                # Calcular largura máxima
                max_length = 0
                column_cells = ws[column_letter]
                
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Definir largura (limitada a 50 para campos muito longos)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar painéis (primeira linha)
        if congelar_paineis:
            ws.freeze_panes = 'A2'
        
        # Adicionar filtros automáticos
        if incluir_filtros:
            ws.auto_filter.ref = ws.dimensions
        
        # Salvar
        wb.save(caminho)
        print(f"[OK] Arquivo Excel formatado salvo em: {caminho}")
        
    except ImportError:
        # Fallback: salvar sem formatação
        print("[AVISO] openpyxl nao instalado. Salvando sem formatacao...")
        df.to_excel(caminho_saida, sheet_name=nome_planilha, index=False, engine='xlsxwriter')
        print(f"[OK] Arquivo Excel salvo em: {caminho_saida}")


def exportar_multiplas_planilhas(
    dfs: dict,
    caminho_saida: str,
    formatar: bool = True
) -> None:
    """
    Exporta múltiplos DataFrames em planilhas separadas do mesmo arquivo.
    
    Args:
        dfs: Dicionário {nome_planilha: DataFrame}
        caminho_saida: Caminho do arquivo .xlsx
        formatar: Se True, aplica formatação
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        caminho = Path(caminho_saida)
        caminho.parent.mkdir(parents=True, exist_ok=True)
        
        # Sanitizar todos os DataFrames antes de salvar
        dfs_limpos = {}
        for nome_planilha, df in dfs.items():
            df_limpo = df.copy()
            for col in df_limpo.columns:
                if df_limpo[col].dtype == 'object':
                    df_limpo[col] = df_limpo[col].apply(sanitizar_para_excel)
            dfs_limpos[nome_planilha] = df_limpo
        
        # Salvar todos os DataFrames
        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            for nome_planilha, df in dfs_limpos.items():
                df.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        if not formatar:
            print(f"[OK] Arquivo Excel com {len(dfs)} planilhas salvo em: {caminho}")
            return
        
        # Carregar e formatar
        wb = load_workbook(caminho)
        
        for nome_planilha in dfs.keys():
            ws = wb[nome_planilha]
            df = dfs[nome_planilha]
            
            # Estilo do cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Cabeçalho
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Dados
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Ajustar larguras
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Congelar e filtrar
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
        
        wb.save(caminho)
        print(f"[OK] Arquivo Excel formatado com {len(dfs)} planilhas salvo em: {caminho}")
        
    except ImportError:
        # Fallback sem formatação
        with pd.ExcelWriter(caminho_saida, engine='xlsxwriter') as writer:
            for nome_planilha, df in dfs.items():
                df.to_excel(writer, sheet_name=nome_planilha, index=False)
        print(f"[OK] Arquivo Excel com {len(dfs)} planilhas salvo em: {caminho_saida}")


def criar_relatorio_completo(
    df_principal: pd.DataFrame,
    caminho_saida: str,
    incluir_estatisticas: bool = True
) -> None:
    """
    Cria um relatório Excel completo com múltiplas planilhas.
    
    Args:
        df_principal: DataFrame principal
        caminho_saida: Caminho do arquivo .xlsx
        incluir_estatisticas: Se True, inclui planilha com estatísticas
    """
    dfs = {"Dados Completos": df_principal}
    
    # Planilha de estatísticas
    if incluir_estatisticas and 'classificacao_final' in df_principal.columns:
        stats_data = {
            'Métrica': [
                'Total de Registros',
                'Desaparecidos sem Desfecho',
                'Desaparecidos Localizados Vivos',
                'Desaparecidos Encontrados Mortos',
                'Desaparecidos Vítimas de Homicídio',
                'Transtornos Psiquiátricos Detectados',
                'Matches Fortes',
                'Matches Moderados',
                'Matches Fracos'
            ],
            'Valor': [
                len(df_principal),
                len(df_principal[df_principal['classificacao_final'] == 'Desaparecido sem desfecho']),
                len(df_principal[df_principal['classificacao_final'] == 'Desaparecido localizado vivo']),
                len(df_principal[df_principal['classificacao_final'] == 'Desaparecido encontrado morto']),
                len(df_principal[df_principal['classificacao_final'] == 'Desaparecido vítima de homicídio']),
                df_principal.get('tem_transtorno_psiquiatrico', pd.Series([False])).sum(),
                df_principal.get('match_forte_cad', pd.Series([False])).sum() + df_principal.get('match_forte_hom', pd.Series([False])).sum(),
                df_principal.get('match_moderado_cad', pd.Series([False])).sum() + df_principal.get('match_moderado_hom', pd.Series([False])).sum(),
                df_principal.get('match_fraco_cad', pd.Series([False])).sum() + df_principal.get('match_fraco_hom', pd.Series([False])).sum()
            ]
        }
        dfs["Estatísticas"] = pd.DataFrame(stats_data)
    
    # Planilha com transtornos
    if 'tem_transtorno_psiquiatrico' in df_principal.columns:
        df_transtornos = df_principal[df_principal['tem_transtorno_psiquiatrico'] == True].copy()
        if len(df_transtornos) > 0:
            colunas_transtorno = ['nome', 'classificacao_final', 'tipo_transtorno', 
                                 'confianca_transtorno', 'evidencia_transtorno']
            colunas_existentes = [col for col in colunas_transtorno if col in df_transtornos.columns]
            dfs["Transtornos Detectados"] = df_transtornos[colunas_existentes]
    
    # Planilha com correlações
    if 'fonte_match' in df_principal.columns:
        df_matches = df_principal[df_principal['fonte_match'].notna()].copy()
        if len(df_matches) > 0:
            colunas_match = ['nome', 'classificacao_final', 'data_desaparecimento', 
                            'data_localizacao_cadaver', 'data_homicidio', 'fonte_match']
            colunas_existentes = [col for col in colunas_match if col in df_matches.columns]
            dfs["Correlações"] = df_matches[colunas_existentes]
    
    # Exportar tudo
    exportar_multiplas_planilhas(dfs, caminho_saida, formatar=True)
